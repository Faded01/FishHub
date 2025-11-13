import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import pandas as pd


class ExcelExporter:
    @staticmethod
    def export_table_to_excel(data, columns, filename=None, sheet_name="Данные"):
        try:
            if not data:
                return False, "Нет данных для экспорта"

            if not columns:
                return False, "Не указаны названия колонок"

            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active

            safe_sheet_name = ExcelExporter._create_safe_sheet_name(sheet_name)
            ws.title = safe_sheet_name

            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell_font = Font(size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col_num, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=str(header))
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border_style
                cell.fill = header_fill

            for row_num, row_data in enumerate(data, 2):
                for col_num, value in enumerate(row_data, 1):
                    if value is None:
                        safe_value = ""
                    elif isinstance(value, (int, float)):
                        safe_value = value
                    else:
                        safe_value = str(value)

                    cell = ws.cell(row=row_num, column=col_num, value=safe_value)
                    cell.font = cell_font
                    cell.alignment = center_alignment
                    cell.border = border_style

            ExcelExporter._adjust_column_widths(ws, columns, data)

            wb.save(filename)
            return True, f"Данные успешно экспортированы в файл: {filename}"

        except Exception as e:
            return False, f"Ошибка экспорта в Excel: {str(e)}"

    @staticmethod
    def _create_safe_sheet_name(name):
        if not name:
            return "Данные"

        forbidden_chars = ['\\', '/', '*', '?', ':', '[', ']']
        safe_name = name
        for char in forbidden_chars:
            safe_name = safe_name.replace(char, '_')

        if len(safe_name) > 31:
            safe_name = safe_name[:31]

        if not safe_name.strip():
            safe_name = "Данные"

        return safe_name

    @staticmethod
    def export_dataframe_to_excel(dataframes, filename=None):
        try:
            if not dataframes:
                return False, "Нет данных для экспорта"

            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_multiple_{timestamp}.xlsx"

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    worksheet = writer.sheets[sheet_name]
                    ExcelExporter._format_excel_sheet(worksheet, df.columns.tolist(), df.values.tolist())

            return True, f"Данные успешно экспортированы в файл: {filename}"

        except Exception as e:
            return False, f"Ошибка экспорта DataFrame в Excel: {str(e)}"

    @staticmethod
    def export_query_results(db_manager, query, params=None, filename=None, columns=None, sheet_name="Данные"):
        try:
            if params:
                db_manager.cursor.execute(query, params)
            else:
                db_manager.cursor.execute(query)

            data = db_manager.cursor.fetchall()

            if not data:
                return False, "Нет данных для экспорта"

            if not columns:
                columns = [description[0] for description in db_manager.cursor.description]

            return ExcelExporter.export_table_to_excel(data, columns, filename, sheet_name)

        except Exception as e:
            return False, f"Ошибка экспорта запроса в Excel: {str(e)}"

    @staticmethod
    def _adjust_column_widths(worksheet, columns, data):
        for col_num, header in enumerate(columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            if header:
                max_length = len(str(header))

            for row_num in range(2, len(data) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _format_excel_sheet(worksheet, columns, data):
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell_font = Font(size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style
            cell.fill = header_fill

        for row_num in range(2, len(data) + 2):
            for col_num in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = cell_font
                cell.alignment = center_alignment
                cell.border = border_style

    @staticmethod
    def get_available_tables_export(db_manager):
        try:
            tables = db_manager.get_table_names()
            return [table for table in tables if not table.startswith('sqlite_')]
        except:
            return []

    @staticmethod
    def export_database_schema(db_manager, filename=None):
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"database_schema_{timestamp}.xlsx"

            tables = ExcelExporter.get_available_tables_export(db_manager)

            wb = openpyxl.Workbook()

            wb.remove(wb.active)

            for table in tables:
                ws = wb.create_sheet(title=table[:31])

                db_manager.cursor.execute(f"PRAGMA table_info({table})")
                columns_info = db_manager.cursor.fetchall()

                schema_headers = ["Имя колонки", "Тип данных", "Обязательное", "Первичный ключ", "По умолчанию"]
                data = []

                for col_info in columns_info:
                    data.append([
                        col_info[1],
                        col_info[2],
                        "Да" if col_info[3] else "Нет",
                        "Да" if col_info[5] else "Нет",
                        col_info[4] if col_info[4] else ""
                    ])


                ExcelExporter.export_table_to_excel(
                    data=data,
                    columns=schema_headers,
                    filename=filename,
                    sheet_name=table[:31]
                )

                source_wb = openpyxl.load_workbook(filename)
                source_ws = source_wb.active

                for row in source_ws.iter_rows():
                    for cell in row:
                        ws[cell.coordinate].value = cell.value
                        ws[cell.coordinate].font = cell.font
                        ws[cell.coordinate].alignment = cell.alignment
                        ws[cell.coordinate].border = cell.border
                        if hasattr(cell, 'fill'):
                            ws[cell.coordinate].fill = cell.fill

                ExcelExporter._adjust_column_widths(ws, schema_headers, data)

            wb.save(filename)
            return True, f"Схема базы данных экспортирована в файл: {filename}"

        except Exception as e:
            return False, f"Ошибка экспорта схемы БД: {str(e)}"